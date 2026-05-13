"""Extract clean tour data from the two Excel files for embedding in index.html."""
import json
import openpyxl
import sys
import re
import pathlib

sys.stdout.reconfigure(encoding="utf-8")

ROOT = pathlib.Path(__file__).parent
TG = pathlib.Path(r"C:\Users\elbics\Downloads\Telegram Desktop")
MARCH = TG / "Тур для Адемы v3.xlsx"
JULY = TG / "Адема Июль 2026 (draft ver_5) (3).xlsx"


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "#NAME?", "#VALUE!"):
        return None
    return s


def parse_excel_time(v):
    """Excel stores time as fractional day. Convert 0.5 → '12:00'."""
    if v is None:
        return None
    try:
        f = float(v)
        if 0 < f < 1:
            total_seconds = round(f * 86400)
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            return f"{h:02d}:{m:02d}"
    except (TypeError, ValueError):
        pass
    return v


def num(v):
    if v is None:
        return None
    try:
        f = float(v)
        return round(f, 2)
    except (TypeError, ValueError):
        return None


def get_rows(path, sheet):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet]
    return list(ws.iter_rows(values_only=True))


# Cache for hyperlink-aware sheets — returns dict {(row, col): (display, hyperlink_target)}
_HL_CACHE = {}


def get_hyperlinks(path, sheet):
    """Return mapping of (row_idx, col_idx) → hyperlink.target string. 1-based indices."""
    key = (path, sheet)
    if key in _HL_CACHE:
        return _HL_CACHE[key]
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[sheet]
    out = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.hyperlink and cell.hyperlink.target:
                out[(cell.row, cell.column)] = cell.hyperlink.target
    _HL_CACHE[key] = out
    return out


def url_or_display(path, sheet, row_idx, col_idx, display_value):
    """Return hyperlink target if exists; else display_value if it looks like URL; else None."""
    hl = get_hyperlinks(path, sheet)
    target = hl.get((row_idx, col_idx))
    if target:
        return target
    if isinstance(display_value, str) and display_value.startswith("http"):
        return display_value
    return None


def parse_march():
    rows = get_rows(MARCH, "Программа")
    days = []
    cur_day = None
    current_date = None
    # row index in iter_rows is 1-based for openpyxl cells (row 1 = first row)
    for idx, r in enumerate(rows[1:], start=2):  # start=2 because we skipped row 1 header
        cleaned = [clean(c) for c in r]
        if not cleaned or len(cleaned) < 11:
            continue
        date_val = cleaned[1] if len(cleaned) > 1 else None
        schedule = cleaned[2] if len(cleaned) > 2 else None
        time_val = parse_excel_time(cleaned[3] if len(cleaned) > 3 else None)
        map_display = cleaned[4] if len(cleaned) > 4 else None
        t_cost = num(cleaned[5] if len(cleaned) > 5 else None)
        t_time = cleaned[6] if len(cleaned) > 6 else None
        t_type = cleaned[7] if len(cleaned) > 7 else None
        hours = cleaned[8] if len(cleaned) > 8 else None
        ticket = num(cleaned[9] if len(cleaned) > 9 else None)
        site_display = cleaned[10] if len(cleaned) > 10 else None

        # Extract real URLs via hyperlinks
        map_url = url_or_display(MARCH, "Программа", idx, 5, map_display)  # col E=5
        site_url = url_or_display(MARCH, "Программа", idx, 11, site_display)  # col K=11

        def mk_event(name, **extra):
            return {
                "name": name,
                "time": time_val,
                "map": map_url,
                "transport_cost": t_cost,
                "transport_time": t_time,
                "transport_type": t_type,
                "hours": hours,
                "ticket": ticket,
                "site": site_url,
                "site_label": site_display if site_display and not (isinstance(site_display, str) and site_display.startswith("http")) else None,
                **extra,
            }

        if date_val and ("Марта" in date_val or "Апреля" in date_val):
            current_date = date_val
            cur_day = {"date": current_date, "events": []}
            days.append(cur_day)
            if schedule:
                cur_day["events"].append(mk_event(schedule))
        elif date_val and ("28.03 бронь" in date_val or "На месте" in date_val or "17:00-17:30" in date_val or "Аниме место" in date_val):
            if cur_day and schedule:
                cur_day["events"].append(mk_event(schedule, tag=date_val))
        elif schedule and cur_day is not None:
            cur_day["events"].append(mk_event(schedule))

    # hotels
    hotel_rows = get_rows(MARCH, "Отели")
    hotels = []
    cur_hotel_section = None
    for idx, r in enumerate(hotel_rows, start=1):
        cleaned = [clean(c) for c in r]
        if not cleaned:
            continue
        first = cleaned[1] if len(cleaned) > 1 else None
        second_display = cleaned[2] if len(cleaned) > 2 else None
        third = cleaned[3] if len(cleaned) > 3 else None
        fourth = cleaned[4] if len(cleaned) > 4 else None
        # Section header detection
        if first and ("с " in first and ("марта" in first or "апреля" in first)):
            cur_hotel_section = first
            continue
        if first == "Отель":
            continue
        # Extract real booking URL
        booking_url = url_or_display(MARCH, "Отели", idx, 3, second_display)
        # Hotel entry (main) — name + link + price + meal
        if first and third and "Стоимость" not in first:
            hotels.append({
                "city_range": cur_hotel_section,
                "name": first,
                "link": booking_url,
                "link_text": second_display if second_display and not (isinstance(second_display, str) and second_display.startswith("http")) else None,
                "price": num(third),
                "meal": fourth,
                "note": cleaned[5] if len(cleaned) > 5 else None,
            })
        # Alternative hotel — only URL (no name) — append as alt for current section
        elif not first and booking_url and cur_hotel_section:
            # Extract a friendly name from URL path
            import urllib.parse
            try:
                path_seg = urllib.parse.urlparse(booking_url).path.split("/")[-1].replace(".html", "").replace(".ru", "")
                friendly = path_seg.replace("-", " ").title()
            except Exception:
                friendly = "Альтернативный отель"
            hotels.append({
                "city_range": cur_hotel_section,
                "name": friendly,
                "link": booking_url,
                "is_alt": True,
            })

    return {
        "title": "Тур для Адемы — v3",
        "dates": "16 марта — 13 апреля 2025",
        "summary": {
            "esim": 26,
            "hotels": 1176,
            "transport_jpy": 91520,
            "tickets_jpy": 79290,
            "currency_note": "Транспорт и билеты в иенах; отели и e-sim в долларах",
        },
        "days": days,
        "hotels": hotels,
    }


def parse_july():
    rows = get_rows(JULY, "Маршрут")
    days = []
    cur_day = None
    last_date = None
    last_dow = None
    last_city = None
    date_re = re.compile(r"^\d{1,2}\s+(July|August|Июля|Августа)", re.IGNORECASE)
    for r in rows[2:]:  # skip 2 header rows
        r = [clean(c) for c in r]
        if not r or len(r) < 13:
            continue
        date_val = r[1] if len(r) > 1 else None
        dow = r[2] if len(r) > 2 else None
        city = r[3] if len(r) > 3 else None
        # skip header row if present
        if date_val == "Dates" or city == "City":
            continue
        time_val = r[4] if len(r) > 4 else None
        place = r[5] if len(r) > 5 else None
        map_link = r[6] if len(r) > 6 else None
        t_cost = num(r[7] if len(r) > 7 else None)
        t_time = r[8] if len(r) > 8 else None
        t_type = r[9] if len(r) > 9 else None
        hours = r[10] if len(r) > 10 else None
        ticket = num(r[11] if len(r) > 11 else None)
        site = r[12] if len(r) > 12 else None

        if date_val:
            last_date = date_val
            last_dow = dow if dow else last_dow
            last_city = city if city else last_city
            cur_day = {"date": last_date, "dow": last_dow, "city": last_city, "events": []}
            days.append(cur_day)
        if city and not date_val:
            last_city = city  # city can change mid-day

        if place and cur_day is not None:
            ev = {
                "name": place,
                "time": time_val,
                "city": city if city else last_city,
                "map": map_link if map_link and map_link.startswith("http") else None,
                "transport_cost": t_cost,
                "transport_time": t_time,
                "transport_type": t_type,
                "hours": hours,
                "ticket": ticket,
                "site": site,
            }
            cur_day["events"].append(ev)

    # attractions catalog
    att_rows = get_rows(JULY, "Attractions")
    attractions = []
    cur_section = None
    CITY_NAMES = ("Удзи", "Киото", "Такасима", "Хиконе", "Нагоя", "Сидзуока", "Фудзикавагутико", "Токио", "Чиба")
    for r_idx, r in enumerate(att_rows, start=1):
        r = [clean(c) for c in r]
        if not r:
            continue
        first = r[1] if len(r) > 1 else None
        price = num(r[2] if len(r) > 2 else None)
        comment = r[3] if len(r) > 3 else None
        status = r[4] if len(r) > 4 else None
        # Section header: city name (may have "Цена/Комментарий/Статус" sub-headers in same row)
        if first in CITY_NAMES:
            cur_section = first
            continue
        # Skip sub-column-header rows
        if first in ("Цена", "Комментарий", "Статус"):
            continue
        if first and first != "Total":
            # Find hyperlink in this row (any column 1-6) for attraction site
            att_url = None
            hl = get_hyperlinks(JULY, "Attractions")
            for c in range(1, 7):
                if (r_idx, c) in hl:
                    att_url = hl[(r_idx, c)]
                    break
            attractions.append({
                "city": cur_section,
                "name": first,
                "price": price,
                "comment": comment,
                "status": status,
                "site": att_url,
            })

    # hotels
    hotel_rows = get_rows(JULY, "Отели")
    hotels = []
    cur_hotel_section = None
    for r in hotel_rows:
        r = [clean(c) for c in r]
        if not r:
            continue
        first = r[1] if len(r) > 1 else None
        room = r[2] if len(r) > 2 else None
        price = num(r[3] if len(r) > 3 else None)
        meal = r[4] if len(r) > 4 else None
        link = r[5] if len(r) > 5 else None
        district = r[6] if len(r) > 6 else None
        note = r[7] if len(r) > 7 else None
        # section header like "11-12 июля Кансай"
        if first and ("июля" in first or "августа" in first) and not room and not price:
            cur_hotel_section = first
            continue
        if first == "Отель":
            continue
        if first and price is not None:
            hotels.append({
                "city_range": cur_hotel_section,
                "name": first,
                "room": room,
                "price": price,
                "meal": meal,
                "link": link if link and link.startswith("http") else None,
                "district": district,
                "note": note,
            })

    # flights
    fl_rows = get_rows(JULY, "Авиабилеты")
    flights = []
    if len(fl_rows) >= 3:
        r = [clean(c) for c in fl_rows[2]]
        # ALM-OSA
        flights.append({
            "route": "Алматы → Осака",
            "date": r[0],
            "airline": r[1],
            "duration": r[2],
            "stop": r[3],
            "price": num(r[4]),
            "baggage": r[5],
            "link": r[6] if r[6] and r[6].startswith("http") else None,
        })
        flights.append({
            "route": "Токио → Алматы",
            "date": r[9],
            "airline": r[10],
            "duration": r[11],
            "stop": r[12],
            "price": num(r[13]),
            "baggage": r[14],
            "link": r[15] if r[15] and r[15].startswith("http") else None,
        })

    # directions
    dir_rows = get_rows(JULY, "Направления")
    directions = []
    for r in dir_rows[3:]:
        r = [clean(c) for c in r]
        if not r:
            continue
        name = r[1] if len(r) > 1 else None
        days_count = r[2] if len(r) > 2 else None
        logistics = num(r[3] if len(r) > 3 else None)
        if name and name != "Активный сценарий":
            directions.append({"city": name, "days": days_count, "logistics": logistics})

    return {
        "title": "Адема в Японии — июль 2026",
        "dates": "11 июля — 2 августа 2026",
        "summary": {
            "flights": 910,
            "esim": 25,
            "hotels": 1211,
            "luggage": 40.48,
            "logistics": 363.00,
            "tour_op": 200,
            "tickets": 548.40,
            "total": 3297.87,
            "currency_note": "Все суммы в USD. JPY→USD по курсу 151.93 (Kansai AP)",
        },
        "days": days,
        "attractions": attractions,
        "hotels": hotels,
        "flights": flights,
        "directions": directions,
    }


data = {
    "march": parse_march(),
    "july": parse_july(),
}

out_json = json.dumps(data, ensure_ascii=False, indent=2)
out_min = json.dumps(data, ensure_ascii=False, separators=(",", ":"))

(ROOT / "tour_data.json").write_text(out_json, encoding="utf-8")
(ROOT / "tour_data.min.js").write_text("window.TOURS=" + out_min + ";\n", encoding="utf-8")

print(f"Wrote tour_data.json ({len(out_json):,} chars)")
print(f"Wrote tour_data.min.js ({len(out_min):,} chars)")
print(f"\nMarch tour: {len(data['march']['days'])} days, {sum(len(d['events']) for d in data['march']['days'])} events, {len(data['march']['hotels'])} hotels")
print(f"July tour: {len(data['july']['days'])} days, {sum(len(d['events']) for d in data['july']['days'])} events, {len(data['july']['hotels'])} hotels, {len(data['july']['attractions'])} attractions, {len(data['july']['flights'])} flights")
